import imaplib
import email
import os
import pandas as pd
import smtplib
import json
import re
from pathlib import Path
import time
import openpyxl
import docx
import PyPDF2
import pytesseract
import textract
from pytesseract import image_to_string
from PIL import Image
from io import BytesIO
import spacy
from spacy.lang.en.stop_words import STOP_WORDS as EN_STOP_WORDS
from spacy.lang.fr.stop_words import STOP_WORDS as FR_STOP_WORDS
from langdetect import detect
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Load spaCy's language models for English and French
nlp_en = spacy.load("en_core_web_sm")
nlp_fr = spacy.load("fr_core_news_sm")

import yaml  # To load saved login credentials from a yaml file

with open("C:/Users/khalil/PycharmProjects/pythonProject1/credentials.yml") as f:
    content = f.read()

# from credentials.yml import user name and password
my_credentials = yaml.load(content, Loader=yaml.FullLoader)

# Load the credentials data from yaml file
time_frequency = my_credentials["time_frequency"]
FOLDER_PATH = r'C:/Users/khalil/Downloads/pdf/'
EMAIL_ADDRESS = my_credentials["EMAIL_ADDRESS"]
EMAIL_PASSWORD = my_credentials["EMAIL_PASSWORD"]
IMAP_URL = 'imap.gmail.com'

# File to attach
EXCEL_FILE_PATH = 'Stages_DataSet.xlsx'


def extract_text_from_docx(docx_file):
    doc = docx.Document(docx_file)
    full_text = []
    for para in doc.paragraphs:
        full_text.append(para.text)
    return "\n".join(full_text)


def extract_text_from_pdf(pdf_file):
    with open(pdf_file, "rb") as file:
        pdf_reader = PyPDF2.PdfReader(file)
        full_text = []
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            full_text.append(page.extract_text())
    return "\n".join(full_text)


def extract_text_from_image(image_file):
    image = Image.open(image_file)
    text = image_to_string(image)
    return text


def extract_text_from_excel(excel_file):
    # Open the Excel workbook
    workbook = openpyxl.load_workbook(excel_file)

    # Initialize a variable to store extracted text
    extracted_text = ""

    # Loop through all sheets in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Loop through all cells in the sheet and extract text
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value:
                    extracted_text += str(cell_value) + "\n"

    return extracted_text


# Function to update the "Attachment Text" column with extracted text from attachments

def update_attachment_text(row):
    attachment_path = row["Attachment"]

    # Check if the attachment path is a string, if not, return "No attachment"
    if not isinstance(attachment_path, str):
        return ""

    # Process the attachment text if the path is valid
    if attachment_path.endswith(".docx"):
        attachment_text = extract_text_from_docx(attachment_path)
    elif attachment_path.endswith(".pdf"):
        attachment_text = extract_text_from_pdf(attachment_path)
    elif attachment_path.endswith((".jpg", ".jpeg", ".png", ".bmp", ".gif")):
        attachment_text = extract_text_from_image(attachment_path)
    elif attachment_path.endswith((".xls", ".xlsx")):
        attachment_text = extract_text_from_excel(attachment_path)  # Call the Excel extraction function

    else:
        # For other unsupported formats, use textract
        try:
            attachment_text = textract.process(attachment_path, encoding='utf-8', errors='ignore').decode("utf-8",
                                                                                                          errors="ignore")
        except Exception as e:
            attachment_text = ""

    return attachment_text


# Function Combine text from attachments and body
def Combine_text(row):
    subject = row["Subject"]
    attachment_text = row["Attachment Text"]
    body_text = row["Body"]

    # Combine and clean the extracted text from attachments and body
    processed_text = f"{subject}\n\n{body_text}\n\n{attachment_text}".strip()
    return processed_text


def replace_french_unicode_escapes(input_string):
    # Define a regular expression pattern to match French Unicode escapes.
    french_unicode_pattern = r'\\u([0-9a-fA-F]{4})'

    # Define a function to replace matched escapes with their corresponding characters.
    def replace_unicode(match):
        return chr(int(match.group(1), 16))

    # Use re.sub to find and replace Unicode escapes.
    result = re.sub(french_unicode_pattern, replace_unicode, input_string)

    return result


# Function to perform text preprocessing
def preprocess_text(text):
    text = text.lower()  # lowercase
    text = replace_french_unicode_escapes(text)
    emoji_pattern = re.compile("["
                               u"\U0001F600-\U0001F64F"  # emoticons
                               u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                               u"\U0001F680-\U0001F6FF"  # transport & map symbols
                               u"\U0001F700-\U0001F77F"  # alchemical symbols
                               u"\U0001F780-\U0001F7FF"  # Geometric Shapes Extended
                               u"\U0001F800-\U0001F8FF"  # Supplemental Arrows-C
                               u"\U0001F900-\U0001F9FF"  # Supplemental Symbols and Pictographs
                               u"\U0001FA00-\U0001FA6F"  # Chess Symbols
                               u"\U0001FA70-\U0001FAFF"  # Symbols and Pictographs Extended-A
                               u"\U0001F000-\U0001F0FF"  # Miscellaneous Symbols and Pictographs
                               u"\U00002700-\U000027BF"  # Dingbats
                               u"\U000024C2-\U0001F251"  # Enclosed Characters
                               "]+", flags=re.UNICODE)
    text = emoji_pattern.sub(r'', text)
    # Detect language
    try:
        language = detect(text)
    except:
        language = None

    # Choose the appropriate spaCy model based on the language
    if language == 'en':
        nlp = nlp_en
        stop_words = EN_STOP_WORDS
    elif language == 'fr':
        nlp = nlp_fr
        stop_words = FR_STOP_WORDS
    else:
        # Default to french if language detection fails or language is not supported
        nlp = nlp_fr
        stop_words = EN_STOP_WORDS

    #     # Tokenization, lowercasing, and lemmatization
    #     doc = nlp(text)
    #     words = [token.lemma_.lower() for token in doc]

    #     # Remove stopwords
    #     words = [word for word in words if word not in stop_words]

    #     # Join the preprocessed words back into a single string
    #     preprocessed_text = ' '.join(words)

    return text, language


def extract_information(text):
    emails = re.findall(r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b', text)
    if emails:
        emails = emails[0]
    else:
        emails = ""
    urls = re.findall(r'http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', text)
    phone_numbers = re.findall(
        r'(\+216[-.\s]?\d{2}[-.\s]?\d{3}[-.\s]?\d{3}|\+216\d{8}|\b\d{8}\b|\b\d{2}[-\s]?\d{3}[-\s]?\d{3}\b)', text)
    return emails, urls, phone_numbers


programming_keywords = ["bi", "react native", "react js", "swagger", "intégration d'apis prédéveloppées", ".net", "ai",
                        "iot", "ansible", "aws", "adobe xd", "agile", "anaconda", "analyste de données sénior",
                        "analytics", "android", "apex", "arango db", "archimate", "architecture micro-services",
                        "azure", "bi tools", "bi business intelligence", "bmc", "bpmn",
                        "bases de données relationnelles", "beautiful soup", "big data", "javascript", "html", "css",
                        "business plan", "business intelligence", "c#", "ci/cd", "css", "travailler en équipe",
                        "communiquer efficacement", "chatter bot", "cloud", "collaboration", "communication",
                        "communication verbale et écrite", "problèmes inversés", "conception de logiciels", "odoo",
                        "react", "angular", "vue.js", "architecture des systèmes d'information", "marketing",
                        "créativité", "dax", "devops", "data analysis", "data analyst", "database management",
                        "deeplearning", "design", "digital marketing strategy", "digital marketing",
                        "intégration sociale", "django", "docker", "développement web", "e-commerce", "eclipse",
                        "elaborer des fiches de missions", "des fiches de contrôle", "expressjs",
                        "expérience des stages", "travail en freelance", "figma", "facebook analytics",
                        "bases de données relationnelles", "sql", "firebase", "flutter", "front-end development",
                        "full-stack development", "fullstack developer", "git", "gitlab", "gestion de projet", "gimsi",
                        "google ux design", "google analytics", "grafana", "heroku", "hibernate", "html", "ios",
                        "imagination", "créativité", "intelligence artificielle", "intégration continue", "ionic",
                        "ionic development", "java", "jee", "jira", "javascript", "jenkins", "k8s", "keycloak",
                        "kotlin", "kubernetes", "suivi des clients", "problèmes d’électricité", "c++", "c",
                        "leadership development", "lucene", "mean or mern stack", "machine learning",
                        "maitrise des notions du clean code", "management", "marketing", "marketing digital", "marvel",
                        "maven", "maîtrise des notions de modèle et méta-modèle", "meteorjs", "micro-service",
                        "microsoft azure", "microsoft dynamics 365", "mobile development", "mongodb", "ms excel",
                        "mysql", "méthodologies agiles", "nlp", "nestjs", "node", "nodejs", "node js", "node.js",
                        "numpy", "oaf", "optimisation de référencement (seo)", "oracle bi", "oracle forms",
                        "organisation", "pdf scrapping", "pestel", "pl/sql", "powerbi", "pandas", "planification",
                        "postgresql", "postman", "powerbi", "powerblas business intelligence", "product owner",
                        "product manager", "professional experience", "python", "backend", "django", "qa analyst",
                        "qlik sense", "reactjs", "rest", "rest api", "rest apis", "react native", "react.js",
                        "recommendation engine", "rest api", "rigueur", "résolution des problèmes informatiques",
                        "résolution des problèmes inversés", "machine learning", "deep learning", "seo", "sgbd", "sql",
                        "svn/git", "swot", "scikit-learn", "scrapy", "scrum", "service now", "software development",
                        "spring", "spring boot", "strong programming skills", "symfony", "tableau desktop",
                        "tableau desktop/mspowerbi", "tailwind", "talend", "task automation", "teaching experience",
                        "data science", "automatic number plate recognition", "anpr", "license plate recognition",
                        "computer vision", "data preprocessing", "image enhancement", "noise reduction",
                        "optical character recognition", "ocr", "state-of-the-art", "travail en équipe", "trello",
                        "typescript", "user experience", "vscode", "volunteering opportunities", "web culture",
                        "web scrapping", "web development", "web load", "web services", "xml", "xml publisher", "xp",
                        "analyse de données", "analyse statistique", "architectures micro-services", "communication",
                        "community management", "conception & modélisation", "copyrighting", "application mobile",
                        "travail collaboratif", "gestion de projet", "développement mobile", "eclipse",
                        "esprit d’analyse", "gestion de projet", "git", "hibernate search", "infrastructure",
                        "invision", "javascript", "json", "langage uml", "mathématiques appliquées en statistiques",
                        "mathématiques", "statistiques", "méthodes agile", "méthodologie agile", "notions unix",
                        "outlook", "pandas", "proto.io", "python", "pytorch", "re", "sketch", "sponsoring",
                        "stratégies marketing", "tabula-py", "tensorflow"]


def extract_skills(text):
    # Initialize an empty list to store extracted skills
    skills = []

    # Convert text to lowercase for case-insensitive matching
    text = text.lower()
    text = re.sub('\s+', ' ', text).strip()  # Remove and double spaces
    # Use regular expression to find matches with programming keywords
    for keyword in programming_keywords:
        if re.search(r'\b' + re.escape(keyword) + r'\b', text):
            skills.append(keyword)


def specific_nlp(text):
    # Detect language
    try:
        language = detect(text)
    except:
        language = None
    if language == 'en':
        return nlp_en
    elif language == 'fr':
        return nlp_fr
    else:
        # Default to french if language is not supported
        return nlp_fr


def extract_named_entities(text):
    nlp = specific_nlp(text)
    doc = nlp(text)
    entities = [(ent.text, ent.label_) for ent in doc.ents]
    return entities
    return skills


def check_and_process_emails():
    # try:
    # Connect to the mailbox
    mail = imaplib.IMAP4_SSL(IMAP_URL)
    mail.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
    mail.select('Inbox')

    # Search for emails
    key = '(OR BODY "stage" BODY "internship")'
    type, data = mail.search(None, key)

    id_list = data[0].split()

    msgs = []
    mylist = []

    for num in data[0].split():
        type, data = mail.fetch(num, '(RFC822)')
        # email_message = email.message_from_string(data[0][1].decode('utf-8'))
        # sender_name, sender_email = email.utils.parseaddr(email_message['from'])#my_msg
        msgs.append(data)

    email_list = []  # List to store email information as dictionaries

    for msg in msgs[::-1]:
        for response_part in msg:
            if isinstance(response_part, tuple):
                email_message = email.message_from_bytes(response_part[1])
                sender_name, sender_email = email.utils.parseaddr(email_message['from'])
        email_info = {
            'Date': email_message['date'],
            'SenderName': sender_name,
            'SenderEmail': sender_email,
            'Recipient': email_message['to'],
            'Subject': email_message['subject'],
            'Body': ''
        }
        # downloading attachments
        for part in email_message.walk():
            if part.get_content_type() == 'text/plain':
                email_info['Body'] = part.get_payload(decode=True).decode('utf-8')
            # this part comes from the snipped I don't understand yet...
            if part.get_content_maintype() == 'multipart':
                continue
            if part.get('Content-Disposition') is None:
                continue
            fileName = part.get_filename()
            fileName = fileName.replace('<', '')
            fileName = fileName.replace('>', '')
            fileName = fileName.replace('=?UTF-8?Q?', '')
            fileName = fileName.replace('=', '')
            fileName = fileName.replace('\r\n\t', '')
            fileName = fileName.replace('?', '')
            print(num)
            if num in mylist:
                pass
            else:
                mylist.append(num)
            if bool(fileName):
                sender = email_message['From']
                sender = sender.replace('<', '')
                sender = sender.replace('>', '')
                sender = sender.replace('=?UTF-8?Q?', '')
                sender = sender.replace('=', '')
                sender = sender.replace('?', '')
                tpath = FOLDER_PATH + str(sender)
                Path(tpath).mkdir(parents=True, exist_ok=True)
                filePath = os.path.join(tpath, fileName)
                # Attempt to open the file, and handle any FileNotFoundError
                if not os.path.isfile(filePath):
                    fp = open(filePath, 'wb')
                    fp.write(part.get_payload(decode=True))
                    fp.close()
                email_info['Attachment'] = filePath
                subject = str(email_message).split("Subject: ", 1)[1].split("\nTo:", 1)[0]
                print('Downloaded "{file}" from email titled "{subject}".'.format(file=fileName, subject=subject))
        email_list.append(email_info)
    for i in mylist:
        mail.copy(i, 'INBOX.Processed')
        mail.store(i, '+FLAGS', '\\Deleted')
    mail.expunge()
    json_file_path = 'emails_with_attachments_data.json'

    # Write the email_list to a JSON file
    with open(json_file_path, 'w', encoding='utf-8') as jsonfile:
        json.dump(email_list, jsonfile, ensure_ascii=False, indent=4)

    print("Email data saved to emails_with_attachments_data.json.")

    # ############################################################print("Read data from the JSON file")
    with open(json_file_path, 'r', encoding='utf-8') as jsonfile:
        email_list = json.load(jsonfile)
    # convert it to a DataFrame using pandas,
    df = pd.DataFrame(email_list)
    # Extracting text from docx pdf image and excel
    df["Attachment Text"] = ""
    df["Attachment Text"] = df.apply(update_attachment_text, axis=1)
    df["Text"] = ""
    # Update the "preprocessed_text" column with extracted text from attachments and body
    df["Text"] = df.apply(Combine_text, axis=1)
    df['preprocessed_text'], df['Language'] = zip(*df['Text'].apply(preprocess_text))
    df['emails'] = ""
    df['urls'] = ""
    # Apply information extraction to the 'preprocessed_text' column
    df['emails'], df['urls'], df['phone_numbers'] = zip(*df['preprocessed_text'].apply(extract_information))
    # Apply the extract_skills function to the "preprocessed_text" column
    df['Skills'] = df['preprocessed_text'].apply(extract_skills)
    # df['Company Name'] = df['preprocessed_text'].apply(extract_company_name)
    print(df)
    # Update Excel file
    df.to_excel("Stages_DataSet.xlsx", index=False)
    mail.copy(num, 'INBOX.Processed')
    mail.store(num, '+FLAGS', '\\Deleted')

    # Expunge deleted emails
    mail.expunge()


# except Exception as e:
#     # Handle exceptions, e.g., log errors
#     print(f"Error: {str(e)}")
    pass